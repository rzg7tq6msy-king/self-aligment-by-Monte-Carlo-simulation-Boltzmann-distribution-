# -*- coding: utf-8 -*-
"""
Created on Mon Sep  8 17:18:12 2025

@author: 25015
"""

import numpy as np
import matplotlib.pyplot as plt
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

from shapely.geometry import Polygon,box,Point,MultiPolygon
from shapely import affinity

#-------------------------选金属电极图案-------------------------
from generate_truncated_ring import generate_truncated_ring_L
from generate_truncated_ring import generate_truncated_ring_M2
from generate_truncated_ring import generate_truncated_ring_S

import energy
from Boltzmann_distribution import boltzmann_probabilities

def simulate_alignment(
        N, sheet_name,
        max_iter=20000, step_shift=1, step_rotate=1,
        converge_window=500, converge_threshold=0.05,
        verbose=False
):
    
    #函数外生成一次文件和文件名
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # --------------------------- 自定义 <simulation_log> 存储位置 -------------------------
    filename = f"C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M2/simulation_log.xlsx"
    
    # 判断文件是否存在           
    # 如果文件存在就加载，否则新建
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    # 获取或创建 sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    # 找到当前最后一列
    last_col = ws.max_column + 1  # 新运行数据写在下一列    
    
    # 写表头到新列
    ws.cell(row=1, column=last_col, value="step") #row行号 #column列号
    ws.cell(row=1, column=last_col+1, value="energy")
    
    # ------------------------自定义 -----------------------
    # ------------------------ 基板 ------------------------
    matrix_s1 = generate_truncated_ring_M2(N)

    # ------------------------ 芯片 ------------------------
    matrix_c1 = generate_truncated_ring_S(N)
    
    
    # ---------------- 初始随机旋转和平移 ----------------

    max_trials = 100  # 最多尝试次数，避免死循环
    for _ in range(max_trials):
        angle = np.random.uniform(0, 360)
        dx, dy = np.random.randint(-N // 10, N // 10, size=2)

        # 旋转和平移,支持 matrix_c1 是 list 或单个 Polygon
        if isinstance(matrix_c1, list):
            moved_chip = [affinity.translate(affinity.rotate(p, angle, origin='center'),xoff=dx, yoff=dy) for p in matrix_c1]
        else:
            moved_chip = affinity.translate(affinity.rotate(matrix_c1, angle, origin='center'),
                                        xoff=dx, yoff=dy)
        
        # 计算与基板重叠面积
        # 用 energy 函数计算重叠面积
        overlap = -energy.energy(matrix_s1, moved_chip)

        if overlap > 0:# 初始芯片必须落在基板上
            current_chip = moved_chip
            current_dx, current_dy = dx, dy
            current_angle = angle
            break
    # 如果尝试 max_trials 次都没有重叠，就接受最后一次
    else:
        current_chip = moved_chip
        current_dx, current_dy = dx, dy
        current_angle = angle

    
    # -------芯片初始状态------------
    current_energy = round(energy.energy(matrix_s1, current_chip),0)
    
    # --------面向显示需求保存每一帧数据-------------
    from output_final_frame import flatten_polygons
    frames = [flatten_polygons(current_chip)]#要输出起始和最终的一帧

    angles_record = [angle]
    # ---------------------------------------------
    energy_record = [current_energy] #初始能量
    steps = [1]  # 保存每个 accepted step 的总迭代步数
    
    for i in range(max_iter):
        move_dx = np.random.choice([-step_shift,0,step_shift]) #  x 方向平移 ±1 或 0
        move_dy = np.random.choice([-step_shift,0,step_shift]) #  y 方向平移 ±1 或 0
        rotate_angle = np.random.choice([-step_rotate,0,step_rotate]) #旋转 ±1 度或不旋转
        
        # 新位置的 polygon
        rotated_chip = affinity.rotate(matrix_c1, angle + rotate_angle, origin='center')
        new_chip = affinity.translate(rotated_chip, xoff=current_dx + move_dx, yoff=current_dy + move_dy)

        # 计算能量（你要确保 energy.energy 支持 polygon 输入）
        new_energy = round(energy.energy(matrix_s1, new_chip),0) #energy没有单位（面积）
        prob = 0 # 初始化
        
        #-------------------------------------------接受移动的判据----------------------------------------
        if new_energy <= current_energy:
            accept = True
        else:
            prob = boltzmann_probabilities(new_energy, current_energy)
            accept = np.random.rand() < prob

        if accept:
            current_energy = new_energy
            current_chip = new_chip
            current_dx += move_dx
            current_dy += move_dy
            angle += rotate_angle
                
        steps.append(steps[-1] + 1) #记录当前循环步数
        #-----------------------------------------------------------------------------------
            
        # -------------------------- 面向显示 -----------------------------
        frames.append(flatten_polygons(current_chip))
        
        energy_record.append(current_energy) #记录每一步演化能量变化的矩阵（实际上是个一维数组）
        angles_record.append(angle % 360)
            
        #-----------------------------------------------------------------
        # 判断收敛：最近 converge_window 步 overlap 变化 < 5%
        if len(energy_record) >= converge_window:  #converge_window=500
            recent = energy_record[-converge_window:] #取出最近 500 步的能量值，用于观察这段时间内的能量变化幅度。
            if abs(max(recent)) < 1e-9:
                change = 0
            else:
                change = (max(recent) - min(recent)) / abs(max(recent))#分子 max(recent) - min(recent) 是最近 500 步中能量的最大差值；分母 abs(max(recent)) 把它归一化为百分比，表示波动占最大值的比例。
            if change < converge_threshold:
                print(f"提前收敛: overlap变化 {change*100:.2f}%")
                break
            
        if verbose and i % 50 == 0:
            print(f"Step {i}: energy = {current_energy:.4f}")

        #数据
        temp = [i,current_energy]# step 和 energy
        print (temp)
        
        ws.cell(row=i+2, column=last_col, value=i)           # step
        ws.cell(row=i+2, column=last_col+1, value=current_energy)  # energy
        
    # 循环结束后保存 Excel 文件
    wb.save(filename)

    return energy_record[0],current_energy,matrix_s1,frames[0],frames[-1]

'''
#调用测试
from output_final_frame import plot_frame
N=169
sheet_name="M4TM4"
initial_energy,current_energy, sub, first_frame,last_frame = simulate_alignment(N,sheet_name)
# 第一帧
plot_frame(sub, first_frame, initial_energy, title='First_Frame', save_path='C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M4/')

# 最终帧
plot_frame(sub, last_frame, current_energy, title='Final_Frame', save_path='C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M4/')

'''




