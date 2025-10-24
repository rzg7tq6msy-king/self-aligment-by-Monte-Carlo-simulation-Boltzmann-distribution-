# -*- coding: utf-8 -*-
"""
Created on Sat Sep  6 14:21:01 2025

@author: 25015
"""

import numpy as np
import matplotlib.pyplot as plt
import matplotlib as mpl

from matplotlib.patches import Polygon as MplPolygon
from matplotlib.collections import PatchCollection


from shapely.geometry import Polygon,box,Point,MultiPolygon
from shapely import affinity


#定义输出图案显示字体为宋体
mpl.rcParams['font.family'] = 'Times New Roman'
mpl.rcParams['axes.unicode_minus'] = True  # 正确显示负号

# ---------- 全局字体大小 ----------
plt.rcParams['font.size'] = 14           # 默认字体大小
plt.rcParams['axes.labelsize'] = 24      # 坐标轴标签大小
plt.rcParams['axes.titlesize'] = 24      # 标题大小
plt.rcParams['xtick.labelsize'] = 16     # x轴刻度字体大小
plt.rcParams['ytick.labelsize'] = 16     # y轴刻度字体大小

# ---------- 全局字体加粗 ----------
plt.rcParams['font.weight'] = 'bold'         # 所有文本默认加粗
plt.rcParams['axes.labelweight'] = 'bold'    # 坐标轴标签加粗
plt.rcParams['axes.titleweight'] = 'bold'    # 标题加粗

# ----------自定义芯片和画布尺寸------------
N=169 #chip size
M=10
scale_factor=np.sqrt(M)

# ----------------------select C to S-----------------------

import simulate_alignment_T2T
#import simulate_alignment_TBH 
#import simulate_alignment_R2R
#import simulate_alignment_S2S

# ---------------------- 每次运行的最终态导入到CSV文件里 ---------------

import os
import csv
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

# --------------------------------------------------------------------
#函数外生成一次文件和文件名
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S") #时间印章

# -------------------------- 自定义<Run_log>存储位置 -------------------------------
#filename = f"C:/Users/25015/Documents/simulate/RUN_log_{timestamp}.xlsx"
filename = f"C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M2/RUN_log.xlsx"
#-----自定义<sheet name>-----
sheet_name = "S2M"

#判断文件是否存在           
# 如果文件存在就加载，否则新建
if os.path.exists(filename):
    wb = load_workbook(filename)
else:
    wb = Workbook()
# 获取或创建 sheet
if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
else:
    ws = wb.create_sheet(sheet_name)
# 找到当前最后一列
last_col = ws.max_column + 1  # 新运行数据写在下一列    

# 写表头到新列
ws.cell(row=1, column=last_col, value="Run times") #row行号 #column列号
ws.cell(row=1, column=last_col+1, value=f"{sheet_name}_energy")
# Make sure the file exists and create it if not


#------------------------自定义运行次数---------------------------
Run_times = 100
energies = []

#循环计算
for j in range(Run_times):
    # ------------------------------ 自定义计算哪个对哪个 ---------------------------------
    initial_energy,final_energy, sub, first_frame,last_frame = simulate_alignment_T2T.simulate_alignment(N,sheet_name)
    
    energies.append(final_energy)
    
    #-----------------------------输出最终态的图----------------------------
    from output_final_frame import polygons_to_patches
    from output_final_frame import plot_frame
    # ------------------------------ 自定义图片保存路径 -----------------------------------
    save_dir = f"C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M2/{sheet_name}"

    plot_frame(sub, last_frame, final_energy, title=f'{final_energy:.0f}_{j+1}', save_path=save_dir)
    
    row = [j,final_energy]
    
    ws.cell(row=j+2, column=last_col, value=j)           # step
    ws.cell(row=j+2, column=last_col+1, value=final_energy)  # energy
# 循环结束后保存 Excel 文件
wb.save(filename)
        
# 把能量映射到 0~1 区间
norm_energies = [e / 2950 for e in energies]

percentage = round(sum(e <= -0.9 for e in norm_energies) / len(norm_energies) * 100,2)

print (percentage)

#---------------------------------输出直方图------------------------------------
plt.figure(figsize=(8,5))
#---------------
hist_values, bins, _ = plt.hist(norm_energies, bins=50, color='skyblue', edgecolor='black', range=(-1, 0))
plt.xlabel("Normalized Energy")
plt.ylabel("Frequency")
plt.title(f"Energy Distribution over {Run_times} Simulations")
plt.grid(True, linestyle="--", alpha=0.7)

# 在图注里加percentage 
plt.text(-0.8, max(hist_values)*0.95, f"{percentage}%", fontsize=16, color='red', ha='center')

#plt.tight_layout() # 调整布局避免重叠

#自定义路径
plt.savefig(f"C:/Users/25015/OneDrive/Documents/pattern/simulate/T2T_M2/{sheet_name}_histogram.png")

plt.show(block=True)












